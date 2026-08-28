"""
Compagnon de collecte guidée pour l'étude "position des capteurs / longueur de tube".

Réutilise directement les modules existants de l'application (config, daq_acquisition,
signal_processing, tube_comparator, reference_base_builder) : les résultats produits
sont donc strictement les mêmes que ceux du bouton "Nouveau test" de l'app principale
(Health Index, corrélation, MAE, Zmax, ratio d'énergie, statut ACCEPTÉ/SUSPECT/REJET).

Ce script ne pilote PAS la position mécanique des capteurs (c'est un réglage manuel
sur le montage) : il vous guide dans la matrice de test définie dans le classeur
"Etude_longueur_position_capteurs.xlsx", déclenche l'acquisition/l'évaluation au bon
moment, et écrit chaque résultat dans un CSV directement compatible avec les feuilles
"Collecte - ..." du classeur (colonnes dans le même ordre, prêtes à copier/coller).

Utilisation :
    1. Placez ce fichier dans le même dossier que main.py (racine UltrasonApp).
    2. python etude_position_capteurs.py
    3. Suivez les instructions à l'écran. Le script reprend automatiquement là où
       vous vous étiez arrêté si vous relancez avec le même fichier de sortie
       (les lignes déjà présentes dans le CSV ne sont pas redemandées).
"""
import os
import sys
import csv
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import config as cfgmod
import daq_acquisition as daqmod
import signal_processing as spmod
import tube_comparator as tcmod
import reference_base_builder as rbb

try:
    import pandas as pd
except ImportError:
    print("pandas est requis (déjà utilisé par l'app principale). Installez-le avec :")
    print("  pip install pandas")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Matrice de test — DOIT rester synchronisée avec le classeur Excel de l'étude.
# Si vous changez les positions dans le classeur, changez-les ici à l'identique.
# ---------------------------------------------------------------------------
PHASE1_POSITIONS_PO = [0, 0.79, 1.57, 2.36, 3.15, 3.94, 4.72, 5.51, 6.30]
PHASE2_POSITIONS_PO = [3.94, 4.72, 5.12, 5.51, 5.91, 6.30]
SIDES = ["Gauche", "Droit"]
REPS = [1, 2, 3]

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "resultats_etude")
os.makedirs(OUTPUT_DIR, exist_ok=True)

PHASE1_CSV = os.path.join(OUTPUT_DIR, "collecte_phase1_tubes_sains.csv")
PHASE2_CSV = os.path.join(OUTPUT_DIR, "collecte_phase2_defaut_connu.csv")

PHASE1_HEADER = [
    "ID_essai", "Date", "Operateur", "N_tube", "Longueur_tube_po",
    "Cote", "Position_capteur_po", "Repetition",
    "SNR_acquisition_dB", "Health_Index", "Correlation_pct",
    "Defauts_P5_P95", "Ratio_Defauts_pct", "MAE", "Zmax", "Energie_ratio",
    "Statut_Base_Saine", "Probabilite_IA", "Diagnostic_IA", "Statut_Final",
    "Amplitude_FFT_max",
]
PHASE2_HEADER = [
    "ID_essai", "Date", "Operateur", "N_tube", "Longueur_tube_po",
    "Cote", "Position_capteur_testee_po", "Position_reelle_defaut_po", "Repetition",
    "SNR_acquisition_dB", "Health_Index", "Correlation_pct",
    "Defauts_P5_P95", "Ratio_Defauts_pct", "MAE", "Zmax", "Energie_ratio",
    "Statut_Base_Saine", "Probabilite_IA", "Diagnostic_IA", "Statut_Final",
    "Defaut_Detecte", "Amplitude_FFT_max",
]


def ask(prompt, default=None):
    suffix = f" [{default}]" if default is not None else ""
    val = input(f"{prompt}{suffix} : ").strip()
    return val if val else default


def ask_float(prompt, default=None):
    while True:
        val = ask(prompt, default)
        try:
            return float(val)
        except (TypeError, ValueError):
            print("  -> valeur numérique attendue, réessayez.")


def load_existing_ids(csv_path):
    """Renvoie l'ensemble des (Cote, Position, Repetition) déjà présents dans le CSV,
    pour permettre de reprendre une session interrompue sans redemander ces lignes."""
    done = set()
    if not os.path.exists(csv_path):
        return done
    with open(csv_path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pos_key = row.get("Position_capteur_po") or row.get("Position_capteur_testee_po")
            done.add((row["Cote"], pos_key, row["Repetition"]))
    return done


def append_row(csv_path, header, row_dict):
    file_exists = os.path.exists(csv_path)
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=header)
        if not file_exists:
            writer.writeheader()
        writer.writerow(row_dict)


def pick_reference_base(cfg):
    bases = rbb.list_reference_bases(cfgmod.REF_BASES_DIR)
    if not bases:
        print("Aucune base de référence trouvée dans data/reference_bases/.")
        print("Créez d'abord une base saine avec l'application principale (main.py).")
        sys.exit(1)
    print("\nBases de référence disponibles :")
    for i, meta in enumerate(bases, start=1):
        print(f"  {i}. {meta['nom']}  ({len(meta.get('tubes_utilises', []))} tubes, "
              f"maj {meta.get('date_maj', '?')})")
    while True:
        choice = ask("Numéro de la base à utiliser", "1")
        try:
            idx = int(choice) - 1
            if 0 <= idx < len(bases):
                meta = bases[idx]
                csv_path = os.path.join(meta["_folder"], "base_saine.csv")
                df_ref = rbb.load_reference_base(csv_path)
                return meta, df_ref
        except ValueError:
            pass
        print("  -> choix invalide, réessayez.")


def acquire_one_tube(cfg, mode):
    """Retourne (DATA, fs_r, n_samples_r, tube_df, snr_acq) pour un tube.
    mode = 'daq' (acquisition via daq_acquisition, réelle ou simulée si NI-DAQ absent)
    mode = 'csv' (import d'un fichier déjà exporté par l'app, colonnes FREQ;FFT Real;FFT Imag;FFT Abs)
    """
    if mode == "daq":
        daq = daqmod.DaqController(cfg)
        real = daq.init_daq()
        if not real:
            print("  (mode simulation — NI-DAQ non détecté, données aléatoires générées)")
        DATA = daq.acquire()
        FREQ_R, FFT_SIGNAL = spmod.compute_fft(
            DATA, daq.fs_r_actual, cfg["F_MIN_FFT"], cfg["F_MAX_FFT"], cfg["N_POINTS_FFT"]
        )
        tube_df = spmod.tube_dataframe(FREQ_R, FFT_SIGNAL)
        snr_acq = spmod.compute_snr(DATA, daq.n_samples_r)
        return tube_df, snr_acq
    else:
        while True:
            path = ask("Chemin du fichier CSV du tube (exporté par l'app)")
            if not path:
                print("  -> chemin vide, réessayez.")
                continue
            path = path.strip().strip('"').strip("'")
            if not os.path.exists(path):
                print(f"  -> fichier introuvable : {path}")
                print("     (vérifiez le chemin complet, ex. C:\\dossier\\tube.csv)")
                continue
            try:
                df = pd.read_csv(path, sep=";")
                df.columns = df.columns.astype(str).str.strip()
                if "FREQ" not in df.columns or cfg["PARAMETRE"] not in df.columns:
                    print(f"  -> colonnes trouvées : {list(df.columns)}")
                    print(f"     attendu : 'FREQ' et '{cfg['PARAMETRE']}' — fichier incompatible.")
                    continue
                return df, None
            except Exception as e:
                print(f"  -> erreur de lecture du fichier : {e}")
                continue


def evaluate_current(cfg, df_ref, tube_df):
    FREQ_R = tube_df["FREQ"].values
    ev = tcmod.evaluate_tube(FREQ_R, tube_df[cfg["PARAMETRE"]].values, df_ref, cfg)
    return ev


def amplitude_fft_max(tube_df):
    """Amplitude brute max du spectre FFT — indicateur physique complémentaire au
    Health Index, calculé à partir des mêmes données déjà acquises (aucun capteur
    ni mesure supplémentaire nécessaire)."""
    return round(float(tube_df["FFT Abs"].max()), 4)


def print_eval_summary(ev, snr_acq):
    snr_txt = "N/A" if snr_acq is None else f"{snr_acq:.2f} dB"
    print(f"  SNR : {snr_txt}  |  Health Index : {ev['health_index']} %  |  "
          f"Statut : {ev['statut_final']}")
    print(f"  Corrélation : {ev['correlation']} %  |  MAE : {ev['mae']}  |  "
          f"Zmax : {ev['zmax']}  |  Énergie : {ev['energie_ratio']}")


def run_phase1(cfg, df_ref, operateur, mode):
    done = load_existing_ids(PHASE1_CSV)
    essai_id = sum(1 for _ in open(PHASE1_CSV, encoding="utf-8")) if os.path.exists(PHASE1_CSV) else 1
    total = len(SIDES) * len(PHASE1_POSITIONS_PO) * len(REPS)
    n = 0
    print(f"\n=== Phase 1 — tubes sains : {total} essais au total ===\n")
    for side in SIDES:
        for pos in PHASE1_POSITIONS_PO:
            for rep in REPS:
                n += 1
                key = (side, str(pos), str(rep))
                if key in done:
                    continue
                print(f"[{n}/{total}] Capteur {side} — position {pos} po — répétition {rep}")
                input("  Positionnez le capteur, puis appuyez sur Entrée pour continuer...")
                n_tube = ask("  N° tube")
                longueur = ask_float("  Longueur du tube (po)")
                tube_df, snr_acq = acquire_one_tube(cfg, mode)
                ev = evaluate_current(cfg, df_ref, tube_df)
                amp_max = amplitude_fft_max(tube_df)
                print_eval_summary(ev, snr_acq)
                row = {
                    "ID_essai": essai_id,
                    "Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Operateur": operateur,
                    "N_tube": n_tube,
                    "Longueur_tube_po": longueur,
                    "Cote": side,
                    "Position_capteur_po": pos,
                    "Repetition": rep,
                    "SNR_acquisition_dB": None if snr_acq is None else round(snr_acq, 2),
                    "Health_Index": ev["health_index"],
                    "Correlation_pct": ev["correlation"],
                    "Defauts_P5_P95": ev["nb_defauts"],
                    "Ratio_Defauts_pct": ev["ratio_defauts"],
                    "MAE": ev["mae"],
                    "Zmax": ev["zmax"],
                    "Energie_ratio": ev["energie_ratio"],
                    "Statut_Base_Saine": ev["statut_base"],
                    "Probabilite_IA": ev["probabilite_ia"],
                    "Diagnostic_IA": ev["diagnostic_ia"],
                    "Statut_Final": ev["statut_final"],
                    "Amplitude_FFT_max": amp_max,
                }
                append_row(PHASE1_CSV, PHASE1_HEADER, row)
                essai_id += 1
    print(f"\nPhase 1 terminée. Résultats dans {PHASE1_CSV}")


def run_phase2(cfg, df_ref, operateur, mode):
    done = load_existing_ids(PHASE2_CSV)
    essai_id = sum(1 for _ in open(PHASE2_CSV, encoding="utf-8")) if os.path.exists(PHASE2_CSV) else 1
    total = len(SIDES) * len(PHASE2_POSITIONS_PO) * len(REPS)
    n = 0
    print(f"\n=== Phase 2 — défaut connu : {total} essais au total ===\n")
    for side in SIDES:
        for pos in PHASE2_POSITIONS_PO:
            for rep in REPS:
                n += 1
                key = (side, str(pos), str(rep))
                if key in done:
                    continue
                print(f"[{n}/{total}] Capteur {side} — position ciblée {pos} po — répétition {rep}")
                input("  Positionnez le capteur sur le défaut connu, puis appuyez sur Entrée...")
                n_tube = ask("  N° tube")
                longueur = ask_float("  Longueur du tube (po)")
                pos_reelle = ask_float("  Position réelle du défaut (po)", pos)
                tube_df, snr_acq = acquire_one_tube(cfg, mode)
                ev = evaluate_current(cfg, df_ref, tube_df)
                amp_max = amplitude_fft_max(tube_df)
                print_eval_summary(ev, snr_acq)
                detecte = "Oui" if ev["statut_final"] != "ACCEPTE" else "Non"
                print(f"  -> Défaut détecté (déduit du statut) : {detecte}")
                row = {
                    "ID_essai": essai_id,
                    "Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Operateur": operateur,
                    "N_tube": n_tube,
                    "Longueur_tube_po": longueur,
                    "Cote": side,
                    "Position_capteur_testee_po": pos,
                    "Position_reelle_defaut_po": pos_reelle,
                    "Repetition": rep,
                    "SNR_acquisition_dB": None if snr_acq is None else round(snr_acq, 2),
                    "Health_Index": ev["health_index"],
                    "Correlation_pct": ev["correlation"],
                    "Defauts_P5_P95": ev["nb_defauts"],
                    "Ratio_Defauts_pct": ev["ratio_defauts"],
                    "MAE": ev["mae"],
                    "Zmax": ev["zmax"],
                    "Energie_ratio": ev["energie_ratio"],
                    "Statut_Base_Saine": ev["statut_base"],
                    "Probabilite_IA": ev["probabilite_ia"],
                    "Diagnostic_IA": ev["diagnostic_ia"],
                    "Statut_Final": ev["statut_final"],
                    "Defaut_Detecte": detecte,
                    "Amplitude_FFT_max": amp_max,
                }
                append_row(PHASE2_CSV, PHASE2_HEADER, row)
                essai_id += 1
    print(f"\nPhase 2 terminée. Résultats dans {PHASE2_CSV}")


def export_to_excel():
    try:
        import openpyxl
    except ImportError:
        print("openpyxl est requis pour l'export. Installez-le avec :")
        print("  pip install openpyxl")
        return

    print("\n=== Export des résultats collectés vers le classeur Excel ===\n")
    xlsx_path = ask("Chemin du classeur Excel (Etude_longueur_position_capteurs.xlsx)")
    if not xlsx_path or not os.path.exists(xlsx_path):
        print("Fichier introuvable.")
        return

    field_to_col_1 = {
        "Date": 2, "Operateur": 3, "N_tube": 4, "Longueur_tube_po": 5,
        "SNR_acquisition_dB": 9, "Health_Index": 10, "Correlation_pct": 11,
        "Defauts_P5_P95": 12, "Ratio_Defauts_pct": 13, "MAE": 14, "Zmax": 15,
        "Energie_ratio": 16, "Statut_Base_Saine": 17, "Amplitude_FFT_max": 19,
    }
    match_cols_1 = {"Cote": 6, "Position_capteur_po": 7, "Repetition": 8}

    field_to_col_2 = {
        "Date": 2, "Operateur": 3, "N_tube": 4, "Longueur_tube_po": 5,
        "Position_reelle_defaut_po": 8,
        "SNR_acquisition_dB": 10, "Health_Index": 11, "Correlation_pct": 12,
        "Defauts_P5_P95": 13, "Ratio_Defauts_pct": 14, "MAE": 15, "Zmax": 16,
        "Energie_ratio": 17, "Statut_Base_Saine": 18, "Statut_Final": 19,
        "Amplitude_FFT_max": 23,
    }
    match_cols_2 = {"Cote": 6, "Position_capteur_testee_po": 7, "Repetition": 9}

    def to_float(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    def row_matches(ws, cell_row, match_cols, csv_row):
        for field, col in match_cols.items():
            cell_val = ws.cell(row=cell_row, column=col).value
            csv_val = csv_row.get(field)
            if field.startswith("Position") or field == "Repetition":
                cv, xv = to_float(csv_val), to_float(cell_val)
                if cv is None or xv is None or abs(cv - xv) > 1e-6:
                    return False
            elif str(cell_val).strip() != str(csv_val).strip():
                return False
        return True

    def write_sheet(ws, csv_rows, field_to_col, match_cols):
        written, not_found = 0, []
        max_row = ws.max_row
        for csv_row in csv_rows:
            found = False
            for r in range(4, max_row + 1):
                if row_matches(ws, r, match_cols, csv_row):
                    for field, col in field_to_col.items():
                        val = csv_row.get(field)
                        fv = to_float(val)
                        ws.cell(row=r, column=col, value=fv if fv is not None else (val or None))
                    written += 1
                    found = True
                    break
            if not found:
                not_found.append({k: csv_row.get(k) for k in match_cols})
        return written, not_found

    wb = openpyxl.load_workbook(xlsx_path)

    rows1 = list(csv.DictReader(open(PHASE1_CSV, encoding="utf-8"))) if os.path.exists(PHASE1_CSV) else []
    if rows1:
        n, missed = write_sheet(wb["Collecte - Tubes sains"], rows1, field_to_col_1, match_cols_1)
        print(f"Phase 1 : {n}/{len(rows1)} lignes écrites.")
        if missed:
            print(f"  {len(missed)} ligne(s) sans correspondance : {missed}")

    rows2 = list(csv.DictReader(open(PHASE2_CSV, encoding="utf-8"))) if os.path.exists(PHASE2_CSV) else []
    if rows2:
        n, missed = write_sheet(wb["Collecte - Défaut connu"], rows2, field_to_col_2, match_cols_2)
        print(f"Phase 2 : {n}/{len(rows2)} lignes écrites.")
        if missed:
            print(f"  {len(missed)} ligne(s) sans correspondance : {missed}")

    wb.save(xlsx_path)
    print(f"\nClasseur mis à jour : {os.path.abspath(xlsx_path)}")
    print("Ouvrez-le dans Excel : les formules se recalculent automatiquement.")


def main():
    print("=== Étude position des capteurs / longueur de tube ===")
    print(f"Les résultats seront écrits dans : {OUTPUT_DIR}\n")
    print("1. Lancer une session de collecte (acquisition + évaluation)")
    print("2. Exporter les résultats collectés vers le classeur Excel")
    choice = ask("Choix", "1")

    if choice == "2":
        export_to_excel()
        return

    cfg = cfgmod.load_config()
    operateur = ask("Nom de l'opérateur")
    meta, df_ref = pick_reference_base(cfg)
    print(f"Base sélectionnée : {meta['nom']}")

    print("\nMode d'acquisition :")
    print("  1. DAQ (réelle si NI-DAQ branché, sinon simulée automatiquement)")
    print("  2. Import CSV (tube déjà exporté par l'app principale)")
    mode = "daq" if ask("Choix", "1") == "1" else "csv"

    print("\nPhase à exécuter :")
    print("  1. Phase 1 — tubes sains (caractérisation par position)")
    print("  2. Phase 2 — tubes avec défaut connu (validation de détection)")
    print("  3. Les deux, à la suite")
    phase = ask("Choix", "1")

    if phase in ("1", "3"):
        run_phase1(cfg, df_ref, operateur, mode)
    if phase in ("2", "3"):
        run_phase2(cfg, df_ref, operateur, mode)

    print("\nTerminé. Relancez ce programme et choisissez l'option 2 pour exporter")
    print("ces résultats vers le classeur Excel de l'étude.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrompu. Relancez le script pour reprendre là où vous vous êtes arrêté.")
    except Exception:
        import traceback
        print("\n=== Une erreur inattendue est survenue ===")
        traceback.print_exc()
        input("\nAppuyez sur Entrée pour fermer...")
